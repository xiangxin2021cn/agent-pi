# /// script
# requires-python = ">=3.12"
# dependencies = ["openpyxl>=3.1,<4", "click>=8.3,<9"]
# ///
"""Excel (.xlsx) operations tool.

Commands: read, write, info, add-sheet, export.

Usage:
    uv run xlsx_tool.py COMMAND [OPTIONS]
"""

import csv
import io
import json
import sys
from pathlib import Path

import click
from openpyxl import Workbook, load_workbook
from openpyxl.utils import range_boundaries


DEFAULT_MAX_ROWS = 2000
DEFAULT_MAX_COLS = 200
DEFAULT_MAX_CELL_CHARS = 4096


def _json_serial(obj: object) -> str:
    """JSON serializer that uses ISO 8601 for dates/datetimes."""
    if hasattr(obj, "isoformat"):
        return obj.isoformat()
    return str(obj)


def write_output(text: str, output_path: str | None) -> None:
    """Write text to file or stdout."""
    if output_path:
        Path(output_path).write_text(text, encoding="utf-8")
        click.echo(f"Output written to {output_path}", err=True)
    else:
        click.echo(text)


@click.group()
def cli() -> None:
    """Excel (.xlsx) operations tool."""
    pass


def _truncate_cell_value(value: object, max_cell_chars: int | None) -> object:
    if value is None or max_cell_chars is None:
        return value
    if isinstance(value, str) and len(value) > max_cell_chars:
        return value[:max_cell_chars] + f"... [truncated {len(value) - max_cell_chars} chars]"
    return value


def _read_sheet_data(
    ws,
    cell_range: str | None = None,
    max_rows: int | None = DEFAULT_MAX_ROWS,
    max_cols: int | None = DEFAULT_MAX_COLS,
    max_cell_chars: int | None = DEFAULT_MAX_CELL_CHARS,
) -> tuple[list[list[object]], bool]:
    """Read bounded data from a worksheet, returning rows and truncation status."""
    if cell_range:
        min_col, min_row, raw_max_col, raw_max_row = range_boundaries(cell_range)
    else:
        min_row = 1
        min_col = 1
        raw_max_row = ws.max_row or 1
        raw_max_col = ws.max_column or 1

    max_row = raw_max_row if max_rows is None else min(raw_max_row, min_row + max_rows - 1)
    max_col = raw_max_col if max_cols is None else min(raw_max_col, min_col + max_cols - 1)
    truncated = max_row < raw_max_row or max_col < raw_max_col

    rows = ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)
    return [
        [_truncate_cell_value(cell.value, max_cell_chars) for cell in row]
        for row in rows
    ], truncated


def _warn_if_truncated(sheet_name: str, max_rows: int | None, max_cols: int | None) -> None:
    limits: list[str] = []
    if max_rows is not None:
        limits.append(f"first {max_rows} rows")
    if max_cols is not None:
        limits.append(f"first {max_cols} columns")
    suffix = " and ".join(limits) if limits else "requested range"
    click.echo(
        f"Warning: sheet '{sheet_name}' was truncated to the {suffix}. "
        "Use --range for a smaller exact area or --no-limit for a full read.",
        err=True,
    )


def _validate_limits(max_rows: int, max_cols: int, max_cell_chars: int, no_limit: bool) -> tuple[int | None, int | None, int | None]:
    if no_limit:
        return None, None, None
    if max_rows <= 0:
        raise click.BadParameter("must be greater than 0", param_hint="--max-rows")
    if max_cols <= 0:
        raise click.BadParameter("must be greater than 0", param_hint="--max-cols")
    if max_cell_chars <= 0:
        raise click.BadParameter("must be greater than 0", param_hint="--max-cell-chars")
    return max_rows, max_cols, max_cell_chars


def _build_records(data: list[list[object]]) -> list[dict[str, object]]:
    """Convert row data (with header row) to list of dicts."""
    if not data or len(data) <= 1:
        return []
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(data[0])]
    records = []
    for row_data in data[1:]:
        record: dict[str, object] = {}
        for i, val in enumerate(row_data):
            key = headers[i] if i < len(headers) else f"col_{i}"
            record[key] = val
        records.append(record)
    return records


def _format_data(data: list[list[object]], fmt: str) -> str:
    """Format row data as text, csv, or json."""
    if not data:
        if fmt == "json":
            return "[]"
        elif fmt == "csv":
            return ""
        else:
            return "(empty)"

    if fmt == "json":
        if len(data) > 1:
            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(data[0])]
            records = []
            for row_data in data[1:]:
                record: dict[str, object] = {}
                for i, val in enumerate(row_data):
                    key = headers[i] if i < len(headers) else f"col_{i}"
                    record[key] = val
                records.append(record)
            return json.dumps(records, indent=2, default=_json_serial)
        else:
            # Single row = header only, no data rows
            return "[]"
    elif fmt == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        for row_data in data:
            writer.writerow(row_data)
        return buf.getvalue()
    else:
        lines: list[str] = []
        str_data = [[str(v) if v is not None else "" for v in row] for row in data]
        if str_data:
            max_cols = max(len(row) for row in str_data)
            col_widths = [0] * max_cols
            for row in str_data:
                for i, val in enumerate(row):
                    col_widths[i] = max(col_widths[i], len(val))
            for row in str_data:
                parts = []
                for i, val in enumerate(row):
                    width = col_widths[i] if i < len(col_widths) else 0
                    parts.append(val.ljust(width))
                lines.append("  ".join(parts).rstrip())
        return "\n".join(lines)


@cli.command()
@click.argument("file", type=click.Path(exists=True, dir_okay=False))
@click.option("--sheet", type=str, default=None, help="Sheet name (default: active sheet).")
@click.option("--all-sheets", is_flag=True, default=False, help="Read all sheets in the workbook.")
@click.option("--range", "cell_range", type=str, default=None, help="Cell range, e.g. 'A1:C10'.")
@click.option("--format", "fmt", type=click.Choice(["text", "csv", "json"]), default="text", help="Output format.")
@click.option("--max-rows", type=int, default=DEFAULT_MAX_ROWS, show_default=True, help="Maximum rows to read per sheet unless --no-limit is set.")
@click.option("--max-cols", type=int, default=DEFAULT_MAX_COLS, show_default=True, help="Maximum columns to read per sheet unless --no-limit is set.")
@click.option("--max-cell-chars", type=int, default=DEFAULT_MAX_CELL_CHARS, show_default=True, help="Maximum characters retained per text cell unless --no-limit is set.")
@click.option("--no-limit", is_flag=True, default=False, help="Read the full requested sheet or range. Use carefully on large workbooks.")
@click.option("-o", "--output", type=click.Path(), default=None, help="Write output to file.")
def read(
    file: str,
    sheet: str | None,
    all_sheets: bool,
    cell_range: str | None,
    fmt: str,
    max_rows: int,
    max_cols: int,
    max_cell_chars: int,
    no_limit: bool,
    output: str | None,
) -> None:
    """Read cells, ranges, or entire sheets from an Excel file."""
    try:
        row_limit, col_limit, cell_char_limit = _validate_limits(max_rows, max_cols, max_cell_chars, no_limit)
        wb = load_workbook(file, read_only=True, data_only=True)

        if all_sheets and sheet:
            wb.close()
            click.echo("Error: --all-sheets and --sheet are mutually exclusive.", err=True)
            sys.exit(1)

        if all_sheets:
            # Read all sheets
            if fmt == "json":
                all_data: dict[str, object] = {}
                for name in wb.sheetnames:
                    ws = wb[name]
                    data, truncated = _read_sheet_data(ws, cell_range, row_limit, col_limit, cell_char_limit)
                    if truncated:
                        _warn_if_truncated(name, row_limit, col_limit)
                    all_data[name] = _build_records(data)
                result = json.dumps(all_data, indent=2, default=_json_serial)
            else:
                parts: list[str] = []
                for name in wb.sheetnames:
                    ws = wb[name]
                    data, truncated = _read_sheet_data(ws, cell_range, row_limit, col_limit, cell_char_limit)
                    if truncated:
                        _warn_if_truncated(name, row_limit, col_limit)
                    if fmt == "csv":
                        parts.append(f"# Sheet: {name}")
                    else:
                        parts.append(f"=== Sheet: {name} ===")
                    parts.append(_format_data(data, fmt))
                    parts.append("")
                result = "\n".join(parts)
            wb.close()
            write_output(result, output)
            return

        if sheet:
            if sheet not in wb.sheetnames:
                wb.close()
                click.echo(f"Error: sheet '{sheet}' not found. Available: {', '.join(wb.sheetnames)}", err=True)
                sys.exit(1)
            ws = wb[sheet]
        else:
            ws = wb.active
            if ws is None:
                click.echo("Error: no active sheet found. Use --sheet to specify one.", err=True)
                wb.close()
                sys.exit(1)

        data, truncated = _read_sheet_data(ws, cell_range, row_limit, col_limit, cell_char_limit)
        if truncated:
            _warn_if_truncated(ws.title, row_limit, col_limit)
        result = _format_data(data, fmt)

        wb.close()
        write_output(result, output)
    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)


@cli.command()
@click.argument("file", type=click.Path(dir_okay=False))
@click.option("--sheet", type=str, default=None, help="Sheet name (default: active sheet).")
@click.option("--cell", type=str, required=True, help="Cell reference, e.g. 'A1'.")
@click.option("--value", type=str, required=True, help="Value to write.")
@click.option("--type", "val_type", type=click.Choice(["string", "number", "bool"]), default="string", help="Value type.")
def write(file: str, sheet: str | None, cell: str, value: str, val_type: str) -> None:
    """Write a value to a specific cell in an Excel file.

    Creates the file if it does not exist.
    """
    try:
        file_path = Path(file)
        if file_path.exists():
            wb = load_workbook(file)
        else:
            wb = Workbook()

        if sheet:
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet)
            ws = wb[sheet]
        else:
            ws = wb.active
            if ws is None:
                click.echo("Error: no active sheet found. Use --sheet to specify one.", err=True)
                wb.close()
                sys.exit(1)

        # Convert value type
        converted: object
        if val_type == "number":
            try:
                converted = int(value)
            except ValueError:
                try:
                    converted = float(value)
                except ValueError:
                    click.echo(f"Error: '{value}' is not a valid number.", err=True)
                    wb.close()
                    sys.exit(1)
        elif val_type == "bool":
            converted = value.lower() in ("true", "1", "yes")
        else:
            converted = value

        ws[cell.upper()] = converted
        wb.save(file)
        wb.close()
        click.echo(f"Wrote '{converted}' to {cell.upper()} in {file_path.name}", err=True)
    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)


@cli.command()
@click.argument("file", type=click.Path(exists=True, dir_okay=False))
@click.option("-o", "--output", type=click.Path(), default=None, help="Write output to file.")
def info(file: str, output: str | None) -> None:
    """Show workbook information: sheets, dimensions, cell counts."""
    try:
        wb = load_workbook(file, read_only=True, data_only=True)

        sheets_info = []
        for name in wb.sheetnames:
            ws = wb[name]
            try:
                dimensions = ws.dimensions
            except Exception:
                # ReadOnlyWorksheet may not expose .dimensions in some openpyxl versions.
                dimensions = ws.calculate_dimension() if hasattr(ws, "calculate_dimension") else None

            sheets_info.append({
                "name": name,
                "dimensions": dimensions,
                "min_row": ws.min_row,
                "max_row": ws.max_row,
                "min_column": ws.min_column,
                "max_column": ws.max_column,
            })

        info_dict = {
            "file": str(Path(file).resolve()),
            "sheet_count": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
            "sheets": sheets_info,
        }

        wb.close()
        write_output(json.dumps(info_dict, indent=2, default=_json_serial), output)
    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)


@cli.command("add-sheet")
@click.argument("file", type=click.Path(dir_okay=False))
@click.option("--name", type=str, required=True, help="Name for the new sheet.")
@click.option("--position", type=int, default=None, help="Position index (0-based). Default: append at end.")
def add_sheet(file: str, name: str, position: int | None) -> None:
    """Add a new sheet to an Excel file.

    Creates the file if it does not exist.
    """
    try:
        file_path = Path(file)
        if file_path.exists():
            wb = load_workbook(file)
        else:
            wb = Workbook()
            # Remove default sheet if creating new file
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        if name in wb.sheetnames:
            wb.close()
            click.echo(f"Error: sheet '{name}' already exists.", err=True)
            sys.exit(1)

        if position is not None:
            wb.create_sheet(name, position)
        else:
            wb.create_sheet(name)

        wb.save(file)
        wb.close()
        click.echo(f"Added sheet '{name}' to {file_path.name}", err=True)
    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)


@cli.command()
@click.argument("file", type=click.Path(exists=True, dir_okay=False))
@click.option("--sheet", type=str, default=None, help="Sheet name (default: active sheet).")
@click.option("--all-sheets", is_flag=True, default=False, help="Export all sheets in the workbook.")
@click.option("--format", "fmt", type=click.Choice(["csv", "json"]), default="csv", help="Export format.")
@click.option("-o", "--output", type=click.Path(), default=None, help="Write output to file.")
def export(file: str, sheet: str | None, all_sheets: bool, fmt: str, output: str | None) -> None:
    """Export a sheet as CSV or JSON."""
    try:
        wb = load_workbook(file, read_only=True, data_only=True)

        if all_sheets and sheet:
            wb.close()
            click.echo("Error: --all-sheets and --sheet are mutually exclusive.", err=True)
            sys.exit(1)

        if all_sheets:
            if fmt == "json":
                all_data: dict[str, object] = {}
                for name in wb.sheetnames:
                    ws = wb[name]
                    data, _ = _read_sheet_data(ws, max_rows=None, max_cols=None, max_cell_chars=None)
                    all_data[name] = _build_records(data)
                result = json.dumps(all_data, indent=2, default=_json_serial)
            else:
                parts: list[str] = []
                for name in wb.sheetnames:
                    ws = wb[name]
                    data, _ = _read_sheet_data(ws, max_rows=None, max_cols=None, max_cell_chars=None)
                    parts.append(f"# Sheet: {name}")
                    parts.append(_format_data(data, "csv"))
                    parts.append("")
                result = "\n".join(parts)
            wb.close()
            write_output(result, output)
            return

        if sheet:
            if sheet not in wb.sheetnames:
                wb.close()
                click.echo(f"Error: sheet '{sheet}' not found. Available: {', '.join(wb.sheetnames)}", err=True)
                sys.exit(1)
            ws = wb[sheet]
        else:
            ws = wb.active
            if ws is None:
                click.echo("Error: no active sheet found. Use --sheet to specify one.", err=True)
                wb.close()
                sys.exit(1)

        data, _ = _read_sheet_data(ws, max_rows=None, max_cols=None, max_cell_chars=None)
        result = _format_data(data, fmt)

        wb.close()
        write_output(result, output)
    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)


if __name__ == "__main__":
    cli()
